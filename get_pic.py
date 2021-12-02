#!/usr/bin/env python
# -*- coding: utf-8 -*-

import configparser
import datetime
import os.path
import shutil
import random
import sys
from time import sleep
import zipfile
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import wait

AUTH_URL = 'https://contebank.com/login'
DOWNLOAD_LINK = f'https://contebank.com/downloadAll'


def read_settings():
    config = configparser.ConfigParser()
    if not os.path.isfile('config.ini'):
        log('Не найден файл конфигурации!')
        return False
    config.read("config.ini", encoding='utf-8-sig')

    try:
        global LOGIN
        LOGIN = config["Settings"]["LOGIN"]
    except KeyError:
        log('Не найдено значение для логина. Работа невозможна!')
        return False

    try:
        global PASSWORD
        PASSWORD = config["Settings"]["PASSWORD"]
    except KeyError:
        log('Не найдено значение для пароля. Работа невозможна!')
        return False

    try:
        global BASE_DIR
        BASE_DIR = config["Settings"]["BASE_DIR"]
    except KeyError:
        log('Не найдено значение для папки, куда сохранять данные!')
        return False

    try:
        global DOWNLOAD_PAUSE
        DOWNLOAD_PAUSE = int(config["Settings"]["DOWNLOAD_PAUSE"])
    except KeyError:
        DOWNLOAD_PAUSE = 30

    return True


def initial():
    if not read_settings():
        return False
    try:
        shutil.rmtree(BASE_DIR)
    except Exception as e:
        log(f'Ошибка ввода пароля: {e}')
    os.mkdir(BASE_DIR)
    if not os.path.isdir(BASE_DIR):
        log(f'Не найдена папка для сохранения данных: {BASE_DIR}')
        return False

    return True


def pause(delay, rand=False):
    if rand:
        if random.randint(1, 3) == 1:
            sign = -1
        else:
            sign = 1
        rand_delay = random.randint(1, delay) / 5
        delay = delay + sign * rand_delay
    sleep(delay)


def log(s):
    now = datetime.datetime.now()
    msg = '{}: {}'.format(now.strftime("%Y-%m-%d %H:%M:%S"), s)
    print(msg)


def get_driver():
    options = webdriver.ChromeOptions()
    # if not DEBUG:
    #     options.add_argument("--no-sandbox")
    #     options.add_argument("--headless")
    #     options.add_argument('--disable-dev-shm-usage')
    #     options.add_argument('--remote-debugging-port=9222')

    options.add_argument('--start-maximized')

    # options.add_argument("user-data-dir=User Data\\")

    # 0- Default, 1 - Enabled, 2 - Disabled
    experimental_flags = ['same-site-by-default-cookies@2', 'cookies-without-same-site-must-be-secure@1']
    chrome_local_state_prefs = {'browser.enabled_labs_experiments': experimental_flags}
    options.add_experimental_option('localState', chrome_local_state_prefs)

    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    options.add_experimental_option("prefs", {
        "download.default_directory": BASE_DIR,
        "browser.download.dir": BASE_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing_for_trusted_sources_enabled": False,
        "safebrowsing.enabled": False,
        "browser.download.folderList": 2,
        "browser.download.manager.showWhenStarting": False,
        "browser.helperApps.neverAsk.saveToDisk": "application/x-gzip,image/webp,image/apng,"
                                                  "text/csv,application/pdf,text/plain,application/csv,"
                                                  "application/vnd.ms-excel, application/zip,"
                                                  "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    })

    driver = webdriver.Chrome(options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    ua = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36'
    driver.execute_cdp_cmd('Network.setUserAgentOverride', {"userAgent": ua})
    log('Драйвер загружен')
    return driver


def get_all_from_page(driver):
    try:
        page_source = driver.page_source
    except:
        page_source = 'no'
    try:
        page_text = driver.find_element("body").text
    except:
        page_text = 'no'

    all_page = page_source + page_text
    return all_page


def clear_field(elem):
    try:
        elem.send_keys(Keys.CONTROL, 'a')
        elem.send_keys(Keys.DELETE)
        pause(0.5)
    except Exception as e:
        log(f'Ошибка очистки поля ввода: {e}')
        return False
    return True


def clear_fields(driver):
    for field in ['name', 'sku_mf']:
        inp = driver.find_element(By.NAME, field)
        clear_field(inp)
    select = Select(driver.find_element(By.NAME, 'brand_id'))
    select.select_by_index(0)
    return True


def wait_for_string_at_page(driver, strings_list, max_wait=60):
    for _ in range(max_wait):
        pause(1, True)
        all_page = get_all_from_page(driver)
        for string in strings_list:
            if string in all_page:
                return True
    return False


def open_main_page(driver):
    log(f'Открытие основной страницы...')
    driver.get(AUTH_URL)
    pause(1, True)
    if wait_for_string_at_page(driver, ['Вход']):
        log('Успешно открыта страница авторизации!')
        return True
    else:
        log('Не удалось открыть страницу авторизации!')
        return False


def make_auth(driver):
    all_page = get_all_from_page(driver)
    if 'Выйти' in all_page:
        log('Авторизация была выполнена ранее!')
        return True

    try:
        inp = driver.find_element(By.NAME, 'email')
        clear_field(inp)
        inp.send_keys(LOGIN)
    except Exception as e:
        log(f'Ошибка ввода логина: {e}')
        return False

    pause(1)
    try:
        inp = driver.find_element(By.NAME, 'password')
        clear_field(inp)

        inp.send_keys(PASSWORD)
        inp.send_keys(Keys.ENTER)
    except Exception as e:
        log(f'Ошибка ввода пароля: {e}')
        return False

    if wait_for_string_at_page(driver, ['Выйти']):
        log('Успешная авторизация!')
        return True
    else:
        log('Не удалось выполнить авторизацию!')
        return False


def find(driver, item, action, item_dict):
    clear_fields(driver)
    inp = driver.find_element(By.NAME, action)
    clear_field(inp)
    inp.send_keys(item)
    inp.send_keys(Keys.ENTER)
    pause(1)
    if len(driver.find_elements(By.XPATH, '//*[@id="models_list"]/div[1]/div')) > 1:
        log(len(driver.find_elements(By.XPATH, '//*[@id="models_list"]/div[1]/div')))
        log('Найдено более одного элемента, пытаемся уточнить поиск')
        select = Select(driver.find_element(By.NAME, 'brand_id'))
        select.select_by_index(1)
        pause(2)
        log(f'Вводим тип товара: {item_dict["product"]}')
        # select.select_by_visible_text('conte elegant')
        # print(select.select_by_index(1))
        # driver.implicitly_wait(10)
        # ActionChains(driver).move_to_element(select).perform()
        # select.select_by_index(1)
        # inp.send_keys(item_dict['tm'])
        # inp.send_keys(Keys.ENTER)
        pause(1)
    if len(driver.find_elements(By.XPATH, '//*[@id="models_list"]/div[1]/div')) > 0:
        # pic = driver.find_element(By.XPATH, '//*[@id="models_list"]/div[1]/div')
        # TODO: Скачать все варианты
        pass
    model = driver.find_element_by_xpath('//*[@id="models_list"]/div[1]/div[1]/a').get_attribute('href')
    # ActionChains(driver).move_to_element(pic).click(pic).perform()
    model = model.split('=')[-1]
    link = f'{DOWNLOAD_LINK}/{model}'

    driver.get(link)

    pause(3)
    for files in os.walk(BASE_DIR):
        for file in files[2]:
            print(file)
            if len(file) > 21:
                os.rename(f'{BASE_DIR}/{file}', f'{BASE_DIR}/{item}.zip')

    return True


def find_item(driver, item):
    print(f'Приходит в find_item: {item}')
    if len(item['item'].split()) > 1 and 'fantasy' in item['item'].lower():
        item_temp = item['item'].split()[1]
    else:
        item_temp = item['item']

    if item_temp.isalnum() and item_temp[0].isdigit():
        log(f'{item_temp} Артикул модели')
        act = 'sku_mf'

    elif item_temp.isalnum() and item_temp[0].isalpha():
        log(f'{item_temp} Название')
        act = 'name'
    else:
        log(f'{item_temp} Название')
        act = 'name'

    find(driver, item_temp, act, item)
    return True


def get_pictures(driver, item):
    find_item(driver, item)


def get_target_list(filename):
    result = []
    wb = openpyxl.load_workbook(filename)
    s = wb.active
    for row in range(2, s.max_row + 1):
        res = {
            'tm': s.cell(row, 2).value,
            'product': s.cell(row, 3).value,
            'item': s.cell(row, 1).value
        }
        result.append(res)
    wb.close()
    print(result)
    return result


def get_brands(driver):
    result = []
    select = Select(driver.find_element(By.NAME, 'brand_id'))
    for index, option in enumerate(select.options):
        result.append((select.options[index].get_attribute('innerText')).replace('\n', '').strip())
    print(result)
    return result


def get_types(driver):
    result = []
    select = Select(driver.find_element(By.NAME, 'type'))
    for index, option in enumerate(select.options):
        result.append((select.options[index].get_attribute('innerText')).replace('\n', '').strip())
    print(result)
    return result


def main(filename):

    if not initial():
        sys.exit()

    log('Начало работы...')

    driver = get_driver()

    if not open_main_page(driver):
        sys.exit()

    if not make_auth(driver):
        sys.exit()

    TARGET_LIST = get_target_list(filename)
    global BRANDS
    BRANDS = get_brands(driver)
    global TYPES
    TYPES = get_types(driver)

    for item in TARGET_LIST:
        print(f'Из TARGET_LIST элемент: {item}')
        get_pictures(driver, item)

    driver.close()
    log('Работа по скачиванию выполнена!')

    for files in os.walk(BASE_DIR):
        for file in files[2]:
            # print(file)
            target_path = f'{BASE_DIR}{os.path.splitext(file)[0]}'
            # try:
            #     shutil.rmtree(target_path)
            # except Exception as e:
            #     log(f'Ошибка удаления директории: {e}')
            zipfile.ZipFile(f'{BASE_DIR}{file}').extractall(path=target_path)
            for fs in os.walk(target_path):
                for index, f in enumerate(fs[2]):
                    target_name = f'{os.path.splitext(file)[0]}_{index}.jpg'
                    # print(f, target_name)
                    os.rename(f'{target_path}/{f}', f'{target_path}/{target_name}')
    log('Выполнено!')
    return True


if __name__ == '__main__':
    main('')
